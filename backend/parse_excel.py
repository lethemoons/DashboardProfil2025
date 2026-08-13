import csv
import sys
import re
import math
from pathlib import Path
import openpyxl

PROVINCE_ONLY = {2, 3, 4, 9, 10, 19, 21, 34, 62}

def clean_col_name(name):
    name = str(name).strip()
    try:
        n = float(name)
        if n == int(n):
            name = str(int(n))
    except ValueError:
        pass
    name = name.lower()
    name = re.sub(r'[^\w\s\+/]', '', name)
    name = re.sub(r'[\s/]+', '_', name)
    name = re.sub(r'_+', '_', name)
    return name.strip('_')

def is_column_number_row(row):
    vals = [v for v in row if v is not None and str(v).strip()]
    if not vals:
        return False
    if not all(isinstance(v, (int, float)) for v in vals):
        return False
    ints = [int(v) for v in vals]
    if ints == list(range(1, len(ints) + 1)):
        return True
    return False

def merge_headers(header_rows):
    ncols = max(len(r) for r in header_rows)
    filled = []
    for row in header_rows:
        r = list(row) + [None] * (ncols - len(row))
        last = None
        for i in range(ncols):
            v = r[i]
            if v is not None and str(v).strip() and str(v).strip() != 'nan':
                last = str(v).strip()
            else:
                r[i] = last
        filled.append(r)

    col_names = []
    for col in range(ncols):
        parts = [str(filled[r][col]).strip() for r in range(len(filled))
                 if filled[r][col] is not None and str(filled[r][col]).strip() != 'nan']
        col_names.append(clean_col_name(' '.join(parts)) if parts else f'col_{col}')
    return col_names

def safe_val(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    s = str(v).strip()
    if s in ('#DIV/0!', '-', '', 'nan'):
        return None
    if s.upper() in ('#VALUE!', '#REF!', '#N/A'):
        return None
    return s

KAB_RE = re.compile(r'^(KAB\.\s*|KOTA\s)', re.IGNORECASE)
PROV_RE = re.compile(r'^PROVINSI$', re.IGNORECASE)
FOOT_RE = re.compile(r'^(SUMBER|CATATAN|Sumber|Catatan)')

def main(input_excel, output_csv):
    try:
        wb = openpyxl.load_workbook(input_excel, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)

    combined_rows = []

    for s in range(1, 89):
        sname = str(s)
        if s in PROVINCE_ONLY:
            continue
        
        if sname not in wb.sheetnames:
            continue

        ws = wb[sname]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        first_data_idx = None
        for i, r in enumerate(all_rows):
            b = r[1] if len(r) > 1 else None
            if b is not None and KAB_RE.search(str(b)):
                first_data_idx = i
                break

        if first_data_idx is None:
            continue

        parsed = []
        for r in all_rows[first_data_idx:]:
            if not r or len(r) < 2:
                continue
            b = str(r[1]).strip() if r[1] is not None else ''
            if not b or not KAB_RE.search(b):
                if PROV_RE.search(b):
                    continue
                if FOOT_RE.search(str(r[0] or '')):
                    break
                continue
            cleaned = [safe_val(v) for v in r]
            parsed.append(cleaned)

        if not parsed:
            continue

        parsed = [row for row in parsed if row[1] is not None and KAB_RE.search(str(row[1]))]
        if not parsed:
            continue

        data_width = 2
        for row in parsed:
            for ci in range(len(row) - 1, -1, -1):
                if row[ci] is not None:
                    data_width = max(data_width, ci + 1)
                    break

        for row in parsed:
            del row[data_width:]

        header_raw = []
        for i in range(first_data_idx - 1, max(first_data_idx - 8, 0) - 1, -1):
            r = all_rows[i]
            if is_column_number_row(r):
                continue
            r_trimmed = r[:data_width] if len(r) > data_width else r + (None,) * (data_width - len(r))
            if any(v is not None for v in r_trimmed):
                header_raw.append(r_trimmed)
            if len(header_raw) >= 2:
                break
        header_raw.reverse()

        if not header_raw:
            f = all_rows[first_data_idx - 1] if first_data_idx > 0 else tuple()
            header_raw = [(f + (None,) * data_width)[:data_width]]

        col_names = merge_headers(header_raw)
        while len(col_names) < data_width:
            col_names.append(f'col_{len(col_names)}')
        col_names = col_names[:data_width]
        if col_names:
            col_names[0] = 'no'
        if len(col_names) > 1:
            col_names[1] = 'kabupaten'

        for row in parsed:
            while len(row) < len(col_names):
                row.append(None)

        for row in parsed:
            kab = row[1]
            no_val = row[0]
            for ci in range(2, len(col_names)):
                if row[ci] is not None:
                    combined_rows.append({
                        'table_no': s,
                        'no': no_val if no_val is not None else '',
                        'kabupaten': kab,
                        'metric': col_names[ci],
                        'value': row[ci]
                    })

    wb.close()

    if combined_rows:
        with open(output_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['table_no', 'no', 'kabupaten', 'metric', 'value'])
            writer.writeheader()
            writer.writerows(combined_rows)
        print(f"SUCCESS: Combined {len(combined_rows)} metric rows into {output_csv}")
    else:
        print("ERROR: No data rows found to combine.")
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python parse_excel.py <input_excel> <output_csv>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    main(input_file, output_file)
